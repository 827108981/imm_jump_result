# -*- coding: utf-8 -*-
import hashlib
import io
import json
import logging
import os
import re
import shutil
import socket
import sys
import threading
import time
import uuid
import webbrowser
import zipfile
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, render_template, request, send_file, send_from_directory
from werkzeug.utils import secure_filename


APP_NAME = "免疫跳值排查反馈报告生成工具"
VERSION = "V1.4"
SHEET_NAME = "用服工程师排查反馈表"
TEMPLATE_NAME = "免疫产品_跳值问题用服工程师排查反馈表_V1.0_CN.xlsx"
PACKAGE_FORMAT_VERSION = "frontline-report-v2"
ZIP_MAX_BYTES = 500 * 1024 * 1024

SOURCE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent

EXTERNAL_RESOURCES_DIR = APP_DIR / "resources"
BUNDLED_RESOURCES_DIR = SOURCE_DIR / "resources"
RESOURCES_DIR = EXTERNAL_RESOURCES_DIR if (EXTERNAL_RESOURCES_DIR / TEMPLATE_NAME).exists() else BUNDLED_RESOURCES_DIR
UPLOADS_DIR = APP_DIR / "uploads"
OUTPUT_DIR = APP_DIR / "output"
DRAFTS_DIR = APP_DIR / "drafts"
LOGS_DIR = APP_DIR / "logs"
RTS_UPLOADS_DIR = UPLOADS_DIR / "rts_reviews"
MOBILE_TASKS_DIR = APP_DIR / "mobile_tasks"
MOBILE_CHUNKS_DIR = APP_DIR / "mobile_upload_chunks"
TEMPLATE_PATH = RESOURCES_DIR / TEMPLATE_NAME

ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}
MAX_IMAGE_SIZE = 5 * 1024 * 1024
IMAGE_COMPRESS_TARGET_SIZE = 2 * 1024 * 1024
MAX_IMAGES_PER_FIELD = 5
MOBILE_TASK_TTL_SECONDS = 8 * 60 * 60
MOBILE_CHUNK_SIZE = 512 * 1024

CONCLUSION_OPTIONS = {
    "normal": "正常",
    "abnormal": "异常",
    "handled": "已处理",
    "pending": "待确认",
}
CONCLUSION_CLASS = {
    "正常": "normal",
    "异常": "abnormal",
    "已处理": "handled",
    "待确认": "pending",
}

SHORT_DESCRIPTION_WORDS = {"正常", "无", "无异常", "OK", "ok", "已完成", "符合", "没问题", "良好", "通过"}

REQUIRED_COLUMNS = [
    "步骤",
    "优先级",
    "分类",
    "排查动作",
    "合格指标",
    "是否执行",
    "实测情况记录",
    "原始状态照片",
    "调试或维护后照片",
]

BASIC_FIELDS = [
    {"key": "hospital", "label": "医院名称", "required": True},
    {"key": "model", "label": "设备型号", "required": True},
    {"key": "serial", "label": "设备序列号", "required": True},
    {"key": "software", "label": "软件版本", "required": True},
    {"key": "jump_project", "label": "跳值项目", "required": True},
    {"key": "problem", "label": "问题描述", "required": True},
    {"key": "engineer", "label": "排查工程师", "required": True},
    {"key": "check_date", "label": "排查日期", "required": True},
    {"key": "contact", "label": "联系方式", "required": False},
]


app = Flask(
    __name__,
    template_folder=str(SOURCE_DIR / "templates"),
    static_folder=str(SOURCE_DIR / "static"),
)
app.config["MAX_CONTENT_LENGTH"] = ZIP_MAX_BYTES + 20 * 1024 * 1024


class UserFacingError(Exception):
    def __init__(self, message, status_code=400):
        super().__init__(message)
        self.message = message
        self.status_code = status_code


def setup_runtime():
    for path in [RESOURCES_DIR, UPLOADS_DIR, RTS_UPLOADS_DIR, OUTPUT_DIR, DRAFTS_DIR, MOBILE_TASKS_DIR, MOBILE_CHUNKS_DIR, LOGS_DIR]:
        path.mkdir(parents=True, exist_ok=True)
    handler = logging.FileHandler(str(LOGS_DIR / "app.log"), encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    root_logger = logging.getLogger()
    root_logger.handlers = [handler]
    root_logger.setLevel(logging.INFO)
    logging.basicConfig(
        handlers=[handler],
        level=logging.INFO,
    )
    logging.info("%s %s started", APP_NAME, VERSION)


def normalize_text(value):
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def is_na(value):
    text = normalize_text(value).upper().replace(" ", "")
    return text in {"N/A", "NA", "N／A"}


def format_display_step(raw_step, fallback_index):
    """Return a clean user-facing step label without duplicating 第/步."""
    text = normalize_text(raw_step)
    if text:
        if text.startswith("第") and text.endswith("步"):
            return text
        return "第%s步" % text
    return "第%s项" % fallback_index


def parse_step_number(raw_step):
    text = normalize_text(raw_step)
    match = re.search(r"(\d+)", text)
    return int(match.group(1)) if match else None


def item_label(item):
    return item.get("display_step") or format_display_step(item.get("step"), item.get("index") or "")


def append_lines(existing, addition):
    existing_text = normalize_text(existing)
    addition_text = normalize_text(addition)
    if not addition_text:
        return existing_text
    if not existing_text:
        return addition_text
    return "%s\n%s" % (existing_text, addition_text)


def merge_template_cell(existing, addition):
    existing_text = normalize_text(existing)
    addition_text = normalize_text(addition)
    if not addition_text or is_na(addition_text):
        return existing_text
    if not existing_text or is_na(existing_text):
        return addition_text
    return append_lines(existing_text, addition_text)


def normalize_conclusion(value):
    text = normalize_text(value) or "正常"
    return text if text in CONCLUSION_CLASS else "正常"


def conclusion_class(value):
    return CONCLUSION_CLASS.get(normalize_conclusion(value), "normal")


def is_short_description(value):
    text = normalize_text(value)
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return False
    if compact in SHORT_DESCRIPTION_WORDS:
        return True
    return len(compact) < 8


def safe_name(value, fallback="未命名"):
    text = normalize_text(value) or fallback
    text = re.sub(r'[<>:"/\\|?*\r\n\t]+', "_", text)
    text = re.sub(r"\s+", "_", text).strip("._ ")
    return text[:80] or fallback


def template_required(value):
    return not is_na(value)


def explicit_template_required(value):
    text = normalize_text(value)
    return bool(text and not is_na(text))


def is_step35_photo_required_action(action):
    return normalize_text(action) in {"本底测试", "样本针和试剂针加样重复性测试", "清洗（缓冲）液检测"}


def is_step35_record_only_action(action):
    return normalize_text(action) == "低值质控重复性测试"


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_extension(filename):
    return filename.rsplit(".", 1)[1].lower() if "." in filename else ""


def import_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise UserFacingError("未安装 openpyxl，请先执行 pip install -r requirements.txt。", 500)
    return load_workbook


def read_excel_template():
    if not TEMPLATE_PATH.exists():
        raise UserFacingError("未找到排查反馈表模板，请确认模板文件是否存在。", 404)

    load_workbook = import_openpyxl()
    workbook = load_workbook(str(TEMPLATE_PATH), data_only=True, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise UserFacingError("模板中未找到“用服工程师排查反馈表”sheet 页，请检查模板版本。", 400)

    sheet = workbook[SHEET_NAME]
    header_row = None
    header_map = {}

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        values = [normalize_text(cell) for cell in row]
        if "排查动作" in values and "合格指标" in values:
            header_row = row_index
            header_map = {name: values.index(name) for name in values if name}
            break

    if not header_row:
        raise UserFacingError("模板中未找到表头，请确认包含“排查动作”和“合格指标”列。", 400)

    missing = [name for name in REQUIRED_COLUMNS if name not in header_map]
    if missing:
        raise UserFacingError("模板缺少以下列：" + "、".join(missing), 400)

    items = []
    current_parent_step_number = None
    current_parent_display_step = ""
    current_parent_priority = ""
    current_parent_category = ""
    step35_child_count = 0

    def build_item(record, item_id, index, display_step, sort_order, priority="", category=""):
        action = record["排查动作"]
        item = {
            "id": item_id,
            "index": index,
            "sort_order": sort_order,
            "step": display_step,
            "display_step": display_step,
            "priority": priority or record["优先级"],
            "category": category or record["分类"] or "未分类",
            "action": action,
            "standard": record["合格指标"],
            "execution": "是",
            "record_template": record["实测情况记录"],
            "before_template": record["原始状态照片"],
            "after_template": record["调试或维护后照片"],
            "record_required": template_required(record["实测情况记录"]),
            "before_required": template_required(record["原始状态照片"]),
            "after_required": template_required(record["调试或维护后照片"]),
        }

        if display_step.startswith("第35步-"):
            if is_step35_photo_required_action(action):
                item["record_required"] = False
                item["before_required"] = False
                item["after_required"] = True
            elif is_step35_record_only_action(action):
                item["record_required"] = True
                item["before_required"] = False
                item["after_required"] = False

        return item

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        record = {name: normalize_text(row[header_map[name]]) for name in REQUIRED_COLUMNS}

        # 正式模板最后可能有“备注”行：步骤列有内容，但排查动作为空。
        # 只要排查动作为空，就不是一个有效排查项。
        if not record["排查动作"]:
            continue

        raw_step = record["步骤"]

        if not raw_step and current_parent_step_number == 35:
            step35_child_count += 1
            display_step = "第35步-%s" % step35_child_count
            items.append(
                build_item(
                    record,
                    "item_035_%s" % step35_child_count,
                    35,
                    display_step,
                    3500 + step35_child_count,
                    current_parent_priority,
                    current_parent_category,
                )
            )
            continue

        if not raw_step and items:
            # 非第35步的空步骤行仍作为上一正式步骤的续行合并。
            last_item = items[-1]
            last_item["action"] = append_lines(last_item["action"], record["排查动作"])
            last_item["standard"] = append_lines(last_item["standard"], record["合格指标"])
            last_item["record_template"] = merge_template_cell(last_item["record_template"], record["实测情况记录"])
            last_item["before_template"] = merge_template_cell(last_item["before_template"], record["原始状态照片"])
            last_item["after_template"] = merge_template_cell(last_item["after_template"], record["调试或维护后照片"])
            last_item["record_required"] = last_item["record_required"] or bool(record["实测情况记录"] and not is_na(record["实测情况记录"]))
            last_item["before_required"] = last_item["before_required"] or bool(record["原始状态照片"] and not is_na(record["原始状态照片"]))
            last_item["after_required"] = last_item["after_required"] or bool(record["调试或维护后照片"] and not is_na(record["调试或维护后照片"]))
            continue

        step_number = parse_step_number(raw_step) or (len(items) + 1)
        current_parent_step_number = step_number
        current_parent_display_step = format_display_step(raw_step, step_number)
        current_parent_priority = record["优先级"]
        current_parent_category = record["分类"]

        if step_number == 35:
            step35_child_count = 1
            display_step = "第35步-1"
            sort_order = 3501
        else:
            display_step = current_parent_display_step
            sort_order = step_number * 100

        items.append(
            build_item(
                record,
                "item_%03d" % step_number,
                step_number,
                display_step,
                sort_order,
                current_parent_priority,
                current_parent_category,
            )
        )

    logging.info("template loaded: %s items from %s", len(items), TEMPLATE_PATH)
    return items


def group_items(items):
    return [{"category": "排查步骤（1-36，第35步含4项）", "items": sorted(items, key=lambda item: item.get("sort_order") or item.get("index") or 0)}]


def get_payload_item_map(payload_items):
    return {item.get("id"): item for item in payload_items or []}


def validate_image_list(images, item, field_key, field_label, errors):
    images = images or []
    if len(images) > MAX_IMAGES_PER_FIELD:
            errors.append(
                {
                    "item_id": item["id"],
                    "field": field_key,
                    "message": "%s：%s每个位置最多上传 %s 张照片。"
                    % (item_label(item), item["action"], MAX_IMAGES_PER_FIELD),
                }
        )

    for image in images:
        original_name = image.get("original_name") or image.get("stored_name") or ""
        if not allowed_file(original_name):
            errors.append(
                {
                    "item_id": item["id"],
                    "field": field_key,
                    "message": "%s：%s%s图片格式不支持。"
                    % (item_label(item), item["action"], field_label),
                }
            )
        if int(image.get("size") or 0) > MAX_IMAGE_SIZE:
            errors.append(
                {
                    "item_id": item["id"],
                    "field": field_key,
                    "message": "%s：%s%s图片超过 5MB。"
                    % (item_label(item), item["action"], field_label),
                }
            )
        try:
            resolve_upload_path(image)
        except UserFacingError as exc:
            errors.append({"item_id": item["id"], "field": field_key, "message": "%s：%s" % (field_label, exc.message)})


def validate_submission(base_info, payload_items, template_items):
    errors = []
    warnings = []
    item_map = get_payload_item_map(payload_items)

    for field in BASIC_FIELDS:
        if field["required"] and not normalize_text(base_info.get(field["key"])):
            errors.append({"field": field["key"], "message": "%s为必填，请补充。" % field["label"]})

    completed_items = 0
    required_text_total = 0
    required_text_done = 0
    missing_image_count = 0
    uploaded_image_count = 0
    compressed_image_count = 0
    conclusion_counts = {label: 0 for label in CONCLUSION_CLASS}

    for item in template_items:
        payload_item = item_map.get(item["id"], {})
        measured = normalize_text(payload_item.get("measured_value"))
        conclusion = normalize_conclusion(payload_item.get("conclusion"))
        before_images = payload_item.get("before_images") or []
        after_images = payload_item.get("after_images") or []
        item_done = True

        conclusion_counts[conclusion] = conclusion_counts.get(conclusion, 0) + 1

        if item["record_required"]:
            required_text_total += 1
            if not measured:
                item_done = False
                errors.append(
                    {
                        "item_id": item["id"],
                        "field": "measured_value",
                        "message": "%s：%s，实测情况记录为必填，请补充。"
                        % (item_label(item), item["action"]),
                    }
                )
            elif len(measured) < 2:
                item_done = False
                errors.append(
                    {
                        "item_id": item["id"],
                        "field": "measured_value",
                        "message": "%s：%s，实测情况记录过短，请补充有效描述。"
                        % (item_label(item), item["action"]),
                    }
                )
            else:
                required_text_done += 1
                if is_short_description(measured):
                    warnings.append(
                        {
                            "item_id": item["id"],
                            "message": "%s：%s，描述较简单，建议补充实测值、观察现象或处理结果。"
                            % (item_label(item), item["action"]),
                        }
                    )

        if item["before_required"] and not before_images:
            item_done = False
            missing_image_count += 1
            errors.append(
                {
                    "item_id": item["id"],
                    "field": "before_images",
                    "message": "%s：%s，原始状态照片必上传。" % (item_label(item), item["action"]),
                }
            )

        if item["after_required"] and not after_images:
            item_done = False
            missing_image_count += 1
            errors.append(
                {
                    "item_id": item["id"],
                    "field": "after_images",
                    "message": "%s：%s，调试或维护后照片必上传。" % (item_label(item), item["action"]),
                }
            )

        validate_image_list(before_images, item, "before_images", "原始状态照片", errors)
        validate_image_list(after_images, item, "after_images", "调试或维护后照片", errors)

        all_images = before_images + after_images
        uploaded_image_count += len(all_images)
        compressed_image_count += sum(1 for image in all_images if image.get("compressed"))
        if item_done:
            completed_items += 1

    stats = {
        "total_items": len(template_items),
        "completed_items": completed_items,
        "uncompleted_items": max(len(template_items) - completed_items, 0),
        "required_text_total": required_text_total,
        "required_text_done": required_text_done,
        "missing_image_count": missing_image_count,
        "uploaded_image_count": uploaded_image_count,
        "compressed_image_count": compressed_image_count,
        "normal_count": conclusion_counts.get("正常", 0),
        "abnormal_count": conclusion_counts.get("异常", 0),
        "handled_count": conclusion_counts.get("已处理", 0),
        "pending_count": conclusion_counts.get("待确认", 0),
        "attention_count": conclusion_counts.get("异常", 0) + conclusion_counts.get("已处理", 0) + conclusion_counts.get("待确认", 0),
        "passed": not errors,
    }
    return errors, stats, warnings


def resolve_upload_path(image):
    session_id = safe_name(image.get("session_id"), "")
    stored_name = image.get("stored_name") or ""
    if not session_id or not stored_name:
        raise UserFacingError("图片缓存信息不完整，请重新上传。")

    path = (UPLOADS_DIR / session_id / stored_name).resolve()
    root = UPLOADS_DIR.resolve()
    if root != path and root not in path.parents:
        raise UserFacingError("图片路径异常，请重新上传。")
    if not path.exists():
        raise UserFacingError("图片缓存不存在，请重新上传。")
    return path


def verify_image(path):
    try:
        from PIL import Image
    except ImportError:
        logging.warning("Pillow is not installed; image verification skipped.")
        return
    try:
        with Image.open(str(path)) as img:
            img.verify()
    except Exception:
        raise UserFacingError("图片读取失败，请重新上传。", 400)


def compress_image_to_limit(path, target_size=IMAGE_COMPRESS_TARGET_SIZE, max_size=MAX_IMAGE_SIZE):
    """Compress an uploaded image in-place-like and return (path, size, compressed)."""
    if path.stat().st_size <= max_size:
        return path, path.stat().st_size, False

    try:
        from PIL import Image, ImageOps
    except ImportError:
        raise UserFacingError("图片超过 5MB，且当前环境未安装 Pillow，无法自动压缩。", 500)

    try:
        with Image.open(str(path)) as img:
            img = ImageOps.exif_transpose(img)
            if img.mode not in ("RGB", "L"):
                img = img.convert("RGB")
            elif img.mode == "L":
                img = img.convert("RGB")

            max_side = 2200
            if max(img.size) > max_side:
                img.thumbnail((max_side, max_side))

            compressed_path = path.with_name(path.stem + "_compressed.jpg")
            quality = 85
            while quality >= 45:
                img.save(str(compressed_path), format="JPEG", quality=quality, optimize=True)
                if compressed_path.stat().st_size <= target_size:
                    break
                quality -= 8

            shrink_round = 0
            while compressed_path.stat().st_size > max_size and shrink_round < 6:
                new_size = (max(800, int(img.width * 0.85)), max(800, int(img.height * 0.85)))
                img = img.resize(new_size)
                img.save(str(compressed_path), format="JPEG", quality=60, optimize=True)
                shrink_round += 1

        if compressed_path.stat().st_size > max_size:
            raise UserFacingError("图片自动压缩后仍超过 5MB，请手动压缩后重新上传。")

        if compressed_path != path and path.exists():
            path.unlink()
        return compressed_path, compressed_path.stat().st_size, True
    except UserFacingError:
        raise
    except Exception:
        logging.exception("image compression failed")
        raise UserFacingError("图片自动压缩失败，请更换图片或手动压缩后重新上传。")


def build_attention_items(report_items):
    return [item for item in report_items if normalize_conclusion(item.get("conclusion")) in {"异常", "已处理", "待确认"}]


def make_issue_no(base_info):
    hospital = normalize_text(base_info.get("hospital")) or "医院"
    model = normalize_text(base_info.get("model")) or "机型"
    serial = normalize_text(base_info.get("serial")) or "序列号"
    return "%s*%s*%s_%s" % (hospital, model, serial, datetime.now().strftime("%Y%m%d%H%M%S"))


def build_package_manifest(base_info, issue_no, generated_at, status="final", export_client="pc", stats=None):
    return {
        "format_version": PACKAGE_FORMAT_VERSION,
        "template_name": TEMPLATE_NAME,
        "tool_version": VERSION,
        "app_name": APP_NAME,
        "export_client": export_client,
        "report_status": status,
        "issue_no": issue_no,
        "generated_at": generated_at,
        "zip_max_bytes": ZIP_MAX_BYTES,
        "summary": {
            "hospital": normalize_text(base_info.get("hospital")),
            "model": normalize_text(base_info.get("model")),
            "serial": normalize_text(base_info.get("serial")),
        },
        "stats": stats or {},
    }


def copy_report_images(base_info, template_items, payload_items, images_dir):
    item_map = get_payload_item_map(payload_items)
    hospital = safe_name(base_info.get("hospital"), "医院")
    serial = safe_name(base_info.get("serial"), "设备序列号")
    report_items = []

    for item in template_items:
        payload_item = item_map.get(item["id"], {})
        report_item = dict(item)
        report_item["measured_value"] = normalize_text(payload_item.get("measured_value"))
        report_item["conclusion"] = normalize_conclusion(payload_item.get("conclusion"))
        report_item["conclusion_class"] = conclusion_class(report_item["conclusion"])
        report_item["short_description_warning"] = is_short_description(report_item["measured_value"])
        report_item["before_images"] = []
        report_item["after_images"] = []

        for field_key, label, output_key in [
            ("before_images", "原始状态照片", "before_images"),
            ("after_images", "调试后照片", "after_images"),
        ]:
            for index, image in enumerate(payload_item.get(field_key) or [], 1):
                source = resolve_upload_path(image)
                ext = get_extension(image.get("stored_name") or image.get("original_name")) or "jpg"
                file_name = "%s_%s_%s_%s_%d.%s" % (
                    hospital,
                    serial,
                    safe_name(item_label(item), "步骤"),
                    label,
                    index,
                    ext,
                )
                file_name = safe_name(file_name, "image") + "." + ext if not file_name.endswith("." + ext) else file_name
                target = images_dir / file_name
                shutil.copy2(str(source), str(target))
                report_item[output_key].append(
                    {
                        "path": "images/" + file_name,
                        "original_name": image.get("original_name") or file_name,
                        "size": image.get("size") or source.stat().st_size,
                        "original_size": image.get("original_size") or image.get("size") or source.stat().st_size,
                        "compressed": bool(image.get("compressed")),
                    }
                )
        report_items.append(report_item)
    return report_items


def make_report_slug(base_info, issue_no=None):
    if issue_no:
        return safe_name(issue_no, datetime.now().strftime("%Y%m%d%H%M%S"))
    return safe_name(make_issue_no(base_info), datetime.now().strftime("%Y%m%d%H%M%S"))


def unique_output_dir(slug):
    target = OUTPUT_DIR / slug
    if not target.exists():
        return target
    return OUTPUT_DIR / ("%s_%s" % (slug, datetime.now().strftime("%H%M%S")))


def build_zip(output_dir, zip_path):
    with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as archive:
        for path in output_dir.rglob("*"):
            if path == zip_path or path.is_dir():
                continue
            archive.write(str(path), str(path.relative_to(output_dir)))


def make_public_output_path(path):
    relative = Path(path).resolve().relative_to(OUTPUT_DIR.resolve())
    return "/output-files/" + str(relative).replace("\\", "/")


def now_ts():
    return int(time.time())


def iso_now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hash_token(token):
    return hashlib.sha256((token or "").encode("utf-8")).hexdigest()


def mobile_task_path(task_id):
    task_name = safe_name(task_id, "")
    if not task_name:
        raise UserFacingError("手机采集任务无效。", 400)
    path = (MOBILE_TASKS_DIR / ("%s.json" % task_name)).resolve()
    root = MOBILE_TASKS_DIR.resolve()
    if root != path and root not in path.parents:
        raise UserFacingError("手机采集任务路径异常。", 400)
    return path


def load_mobile_task(task_id):
    path = mobile_task_path(task_id)
    if not path.exists():
        raise UserFacingError("手机采集任务不存在，请在电脑端重新生成二维码。", 404)
    return json.loads(path.read_text(encoding="utf-8"))


def save_mobile_task(task):
    path = mobile_task_path(task.get("task_id"))
    path.write_text(json.dumps(task, ensure_ascii=False, indent=2), encoding="utf-8")


def assert_mobile_task_token(task, token):
    if task.get("status") != "active":
        raise UserFacingError("手机采集任务已关闭，请在电脑端重新生成二维码。", 403)
    if int(task.get("expires_at_ts") or 0) < now_ts():
        task["status"] = "expired"
        save_mobile_task(task)
        raise UserFacingError("二维码已过期，请在电脑端刷新二维码。", 403)
    if hash_token(token) != task.get("token_hash"):
        raise UserFacingError("手机采集口令无效，请重新扫码。", 403)


def find_payload_item(task, item_id):
    for item in task.get("items") or []:
        if item.get("id") == item_id:
            return item
    item = {
        "id": item_id,
        "measured_value": "",
        "conclusion": "正常",
        "before_images": [],
        "after_images": [],
    }
    task.setdefault("items", []).append(item)
    return item


def append_unique_images(item, field, images):
    existing = item.setdefault(field, [])
    existing_names = {image.get("stored_name") for image in existing}
    for image in images:
        if image.get("stored_name") not in existing_names:
            image["source"] = "mobile"
            existing.append(image)
            existing_names.add(image.get("stored_name"))


def mobile_chunk_task_dir(task_id):
    task_name = safe_name(task_id, "")
    if not task_name:
        raise UserFacingError("手机采集任务无效。", 400)
    path = (MOBILE_CHUNKS_DIR / task_name).resolve()
    root = MOBILE_CHUNKS_DIR.resolve()
    if root != path and root not in path.parents:
        raise UserFacingError("分片缓存路径异常。", 400)
    return path


def mobile_chunk_upload_dir(task_id, upload_id):
    upload_name = safe_name(upload_id, "")
    if not upload_name:
        raise UserFacingError("上传任务无效。", 400)
    path = (mobile_chunk_task_dir(task_id) / upload_name).resolve()
    root = mobile_chunk_task_dir(task_id).resolve()
    if root != path and root not in path.parents:
        raise UserFacingError("分片上传路径异常。", 400)
    return path


def chunk_metadata_path(task_id, upload_id):
    return mobile_chunk_upload_dir(task_id, upload_id) / "metadata.json"


def load_chunk_metadata(task_id, upload_id):
    path = chunk_metadata_path(task_id, upload_id)
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_chunk_metadata(task_id, upload_id, metadata):
    upload_dir = mobile_chunk_upload_dir(task_id, upload_id)
    upload_dir.mkdir(parents=True, exist_ok=True)
    (upload_dir / "chunks").mkdir(parents=True, exist_ok=True)
    chunk_metadata_path(task_id, upload_id).write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_positive_int(value, label, minimum=1, maximum=None):
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise UserFacingError("%s无效。" % label, 400)
    if number < minimum:
        raise UserFacingError("%s无效。" % label, 400)
    if maximum is not None and number > maximum:
        raise UserFacingError("%s超出限制。" % label, 400)
    return number


def uploaded_chunk_indexes(task_id, upload_id):
    chunks_dir = mobile_chunk_upload_dir(task_id, upload_id) / "chunks"
    if not chunks_dir.exists():
        return []
    indexes = []
    for path in chunks_dir.glob("*.part"):
        try:
            indexes.append(int(path.stem))
        except ValueError:
            continue
    return sorted(set(indexes))


def validate_mobile_upload_context(task, item_id, field):
    if field not in {"before_images", "after_images"}:
        raise UserFacingError("图片上传位置无效。", 400)
    if not item_id:
        raise UserFacingError("未指定排查项。", 400)
    return find_payload_item(task, item_id)


def get_lan_ipv4_addresses():
    addresses = []
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            addresses.append(sock.getsockname()[0])
    except OSError:
        pass

    try:
        hostname = socket.gethostname()
        for item in socket.getaddrinfo(hostname, None, socket.AF_INET, socket.SOCK_STREAM):
            address = item[4][0]
            if address and not address.startswith("127."):
                addresses.append(address)
    except OSError:
        pass

    unique = []
    for address in addresses:
        if address not in unique and not address.startswith("127."):
            unique.append(address)
    return unique


def get_request_port():
    host = request.host or ""
    if ":" in host:
        return host.rsplit(":", 1)[-1]
    return str(int(os.environ.get("JUMP_CHECK_PORT") or 5000))


def build_mobile_url(task_id, token, preferred_ip=None):
    ip = preferred_ip or (get_lan_ipv4_addresses() or ["127.0.0.1"])[0]
    return "http://%s:%s/mobile?task_id=%s&token=%s" % (ip, get_request_port(), task_id, token)


def create_mobile_task_from_payload(payload):
    session_id = safe_name(payload.get("session_id") or uuid.uuid4().hex, uuid.uuid4().hex)
    task_id = "m_%s_%s" % (datetime.now().strftime("%Y%m%d%H%M%S"), uuid.uuid4().hex[:8])
    token = uuid.uuid4().hex + uuid.uuid4().hex
    mobile_url = build_mobile_url(task_id, token)
    task = {
        "task_id": task_id,
        "session_id": session_id,
        "token_hash": hash_token(token),
        "status": "active",
        "created_at": iso_now(),
        "expires_at": datetime.fromtimestamp(now_ts() + MOBILE_TASK_TTL_SECONDS).strftime("%Y-%m-%d %H:%M:%S"),
        "expires_at_ts": now_ts() + MOBILE_TASK_TTL_SECONDS,
        "last_seen_at": "",
        "updated_at": iso_now(),
        "updated_seq": 0,
        "mobile_url": mobile_url,
        "base_info": payload.get("base_info") or {},
        "items": payload.get("items") or [],
    }
    save_mobile_task(task)
    return task, token


def mobile_task_response(task, token):
    return {
        "ok": True,
        "task_id": task["task_id"],
        "token": token,
        "mobile_url": task["mobile_url"],
        "qr_url": "/api/mobile/task/%s/qrcode" % task["task_id"],
        "lan_ips": get_lan_ipv4_addresses(),
        "expires_at": task["expires_at"],
    }


def save_uploaded_files(session_id, files):
    session_id = safe_name(session_id or uuid.uuid4().hex, uuid.uuid4().hex)
    upload_dir = UPLOADS_DIR / session_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    if not files:
        raise UserFacingError("未选择图片。")

    uploaded = []
    for file_storage in files:
        original_name = file_storage.filename or ""
        if not allowed_file(original_name):
            raise UserFacingError("图片格式不支持，请上传 jpg、jpeg、png 或 webp 格式图片。")

        file_storage.stream.seek(0, os.SEEK_END)
        original_size = file_storage.stream.tell()
        file_storage.stream.seek(0)

        ext = get_extension(original_name)
        base = Path(secure_filename(original_name)).stem or uuid.uuid4().hex
        stored_name = "%s_%s_%s.%s" % (int(time.time() * 1000), uuid.uuid4().hex[:8], safe_name(base, "image"), ext)
        target = upload_dir / stored_name
        file_storage.save(str(target))
        verify_image(target)

        target, final_size, compressed = compress_image_to_limit(target)
        verify_image(target)
        stored_name = target.name

        uploaded.append(
            {
                "session_id": session_id,
                "stored_name": stored_name,
                "original_name": original_name,
                "size": final_size,
                "original_size": original_size,
                "compressed": compressed,
                "url": "/uploads/%s/%s" % (session_id, stored_name),
            }
        )

    logging.info("uploaded %s images to session %s", len(uploaded), session_id)
    return session_id, uploaded


def register_uploaded_image(session_id, source_path, original_name, original_size=None, extra=None):
    session_id = safe_name(session_id or uuid.uuid4().hex, uuid.uuid4().hex)
    upload_dir = UPLOADS_DIR / session_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    original_name = original_name or source_path.name
    if not allowed_file(original_name):
        raise UserFacingError("图片格式不支持，请上传 jpg、jpeg、png 或 webp 格式图片。")

    ext = get_extension(original_name) or get_extension(source_path.name) or "jpg"
    base = Path(secure_filename(original_name)).stem or uuid.uuid4().hex
    stored_name = "%s_%s_%s.%s" % (int(time.time() * 1000), uuid.uuid4().hex[:8], safe_name(base, "image"), ext)
    target = upload_dir / stored_name
    shutil.move(str(source_path), str(target))

    verify_image(target)
    target, final_size, compressed = compress_image_to_limit(target)
    verify_image(target)

    image = {
        "session_id": session_id,
        "stored_name": target.name,
        "original_name": original_name,
        "size": final_size,
        "original_size": original_size or final_size,
        "compressed": compressed,
        "url": "/uploads/%s/%s" % (session_id, target.name),
    }
    if extra:
        image.update(extra)
    return session_id, image


def cache_existing_report_image(session_id, source_path, image_info):
    session_id = safe_name(session_id or uuid.uuid4().hex, uuid.uuid4().hex)
    upload_dir = UPLOADS_DIR / session_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    original_name = image_info.get("original_name") or source_path.name
    ext = get_extension(original_name) or get_extension(source_path.name) or "jpg"
    if ext not in ALLOWED_EXTENSIONS:
        ext = "jpg"
    base = Path(secure_filename(original_name)).stem or Path(source_path).stem or uuid.uuid4().hex
    stored_name = "%s_%s_%s.%s" % (int(time.time() * 1000), uuid.uuid4().hex[:8], safe_name(base, "image"), ext)
    target = upload_dir / stored_name
    shutil.copy2(str(source_path), str(target))

    verify_image(target)
    target, final_size, compressed_now = compress_image_to_limit(target)
    verify_image(target)

    return {
        "session_id": session_id,
        "stored_name": target.name,
        "original_name": original_name,
        "size": final_size,
        "original_size": image_info.get("original_size") or image_info.get("size") or source_path.stat().st_size,
        "compressed": bool(image_info.get("compressed")) or compressed_now,
        "url": "/uploads/%s/%s" % (session_id, target.name),
        "source": "restored_zip",
    }


@app.errorhandler(UserFacingError)
def handle_user_error(exc):
    logging.warning("user-facing error: %s", exc.message)
    return jsonify({"ok": False, "message": exc.message}), exc.status_code


@app.errorhandler(Exception)
def handle_unexpected_error(exc):
    logging.exception("unexpected error")
    return jsonify({"ok": False, "message": "程序异常，请查看 logs/app.log。"}), 500


@app.route("/")
def index():
    return render_template(
        "index.html",
        app_name=APP_NAME,
        version=VERSION,
        basic_fields=BASIC_FIELDS,
        template_name=TEMPLATE_NAME,
    )


def render_mobile_offline_html():
    return (SOURCE_DIR / "templates" / "mobile_offline.html").read_text(encoding="utf-8")


@app.route("/mobile-offline.html")
def mobile_offline_preview():
    return render_mobile_offline_html()


@app.route("/mobile-offline/download")
def mobile_offline_download():
    html = render_mobile_offline_html()
    buffer = io.BytesIO(html.encode("utf-8"))
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="text/html; charset=utf-8",
        as_attachment=True,
        download_name="mobile_offline.html",
    )


@app.route("/api/template")
def api_template():
    items = read_excel_template()
    return jsonify(
        {
            "ok": True,
            "items": items,
            "groups": group_items(items),
            "basic_fields": BASIC_FIELDS,
            "package_format_version": PACKAGE_FORMAT_VERSION,
            "template_name": TEMPLATE_NAME,
            "limits": {
                "allowed_extensions": sorted(ALLOWED_EXTENSIONS),
                "max_image_size": MAX_IMAGE_SIZE,
                "max_images_per_field": MAX_IMAGES_PER_FIELD,
                "zip_max_bytes": ZIP_MAX_BYTES,
            },
        }
    )


@app.route("/api/upload", methods=["POST"])
def api_upload():
    session_id, uploaded = save_uploaded_files(request.form.get("session_id"), request.files.getlist("files"))
    return jsonify({"ok": True, "session_id": session_id, "files": uploaded})


@app.route("/api/check", methods=["POST"])
def api_check():
    payload = request.get_json(force=True)
    template_items = read_excel_template()
    errors, stats, warnings = validate_submission(
        payload.get("base_info") or {},
        payload.get("items") or [],
        template_items,
    )
    if errors:
        logging.info("validation failed: %s errors", len(errors))
    return jsonify({"ok": True, "passed": not errors, "errors": errors, "warnings": warnings, "stats": stats})


@app.route("/api/report", methods=["POST"])
def api_report():
    payload = request.get_json(force=True)
    base_info = payload.get("base_info") or {}
    payload_items = payload.get("items") or []
    template_items = read_excel_template()
    errors, stats, warnings = validate_submission(base_info, payload_items, template_items)

    if errors:
        logging.info("report blocked by validation: %s errors", len(errors))
        return jsonify({"ok": False, "passed": False, "errors": errors, "warnings": warnings, "stats": stats}), 400

    issue_no = make_issue_no(base_info)
    report_base_info = dict(base_info)
    report_base_info["issue_no"] = issue_no
    slug = make_report_slug(base_info, issue_no)
    output_dir = unique_output_dir(slug)
    images_dir = output_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    report_items = copy_report_images(base_info, template_items, payload_items, images_dir)
    report_groups = group_items(report_items)
    attention_items = build_attention_items(report_items)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    manifest = build_package_manifest(report_base_info, issue_no, generated_at, "final", "pc", stats)

    report_html = render_template(
        "report.html",
        app_name=APP_NAME,
        version=VERSION,
        base_info=report_base_info,
        basic_fields=BASIC_FIELDS,
        stats=stats,
        groups=report_groups,
        attention_items=attention_items,
        issue_no=issue_no,
        generated_at=generated_at,
    )

    report_path = output_dir / "report.html"
    manifest_path = output_dir / "manifest.json"
    data_path = output_dir / "report_data.json"
    zip_path = output_dir / (slug + ".zip")

    report_path.write_text(report_html, encoding="utf-8")
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    data_path.write_text(
        json.dumps(
            {
                "format_version": PACKAGE_FORMAT_VERSION,
                "base_info": report_base_info,
                "issue_no": issue_no,
                "stats": stats,
                "warnings": warnings,
                "groups": report_groups,
                "attention_items": attention_items,
                "generated_at": generated_at,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    build_zip(output_dir, zip_path)
    logging.info("report generated: %s", report_path)

    return jsonify(
        {
            "ok": True,
            "passed": True,
            "stats": stats,
            "issue_no": issue_no,
            "output_dir": str(output_dir),
            "report_file": str(report_path),
            "zip_file": str(zip_path),
            "report_url": make_public_output_path(report_path),
            "zip_url": make_public_output_path(zip_path),
        }
    )


@app.route("/api/report/import", methods=["POST"])
def api_report_import():
    file_storage = request.files.get("report_file")
    if not file_storage or not file_storage.filename:
        raise UserFacingError("请选择一线原始报告 ZIP。")
    if not file_storage.filename.lower().endswith(".zip"):
        raise UserFacingError("目前仅支持导入一线生成的 ZIP 报告包。")

    import_session_id = "restore_src_" + uuid.uuid4().hex
    session_root = RTS_UPLOADS_DIR / import_session_id
    session_root.mkdir(parents=True, exist_ok=True)
    zip_path = session_root / "source_report.zip"
    file_storage.save(str(zip_path))

    source_data, extract_root = load_report_data_from_zip(zip_path, import_session_id)
    restored = restore_editable_report_from_source(source_data, extract_root)
    logging.info("frontline report restored: source=%s session=%s items=%s", import_session_id, restored["session_id"], len(restored["items"]))
    return jsonify({"ok": True, **restored})


@app.route("/api/report/import-rts-review", methods=["POST"])
def api_report_import_rts_review():
    file_storage = request.files.get("rts_file")
    if not file_storage or not file_storage.filename:
        raise UserFacingError("请选择 RTS 审核返回 ZIP 或 rts_review_data.json。")

    session_id = "frontline_rts_" + uuid.uuid4().hex
    session_root = RTS_UPLOADS_DIR / session_id
    session_root.mkdir(parents=True, exist_ok=True)
    lower = file_storage.filename.lower()

    if lower.endswith(".zip"):
        zip_path = session_root / "rts_review.zip"
        file_storage.save(str(zip_path))
        rts_data, _ = load_rts_review_data_from_zip(zip_path, session_id)
    elif lower.endswith(".json"):
        rts_data, _ = load_rts_review_data_from_json(file_storage, session_id)
    else:
        raise UserFacingError("文件格式不支持，请上传 RTS 返回 ZIP 或 rts_review_data.json。")

    source_data = rts_data.get("source_data") or {}
    supplement_requests = normalize_supplement_requests(rts_data.get("supplement_requests") or [], source_data)
    logging.info("RTS return imported by frontline: session=%s review=%s requests=%s", session_id, rts_data.get("review_no"), len(supplement_requests))
    return jsonify(
        {
            "ok": True,
            "review_no": rts_data.get("review_no") or "",
            "source_issue_no": rts_data.get("source_issue_no") or "",
            "review": rts_data.get("review") or {},
            "review_tags": rts_data.get("review_tags") or {},
            "supplement_requests": supplement_requests,
            "generated_at": rts_data.get("generated_at") or "",
        }
    )


@app.route("/mobile")
def mobile_page():
    return render_template("mobile.html", app_name=APP_NAME, version=VERSION)


@app.route("/api/mobile/task/create", methods=["POST"])
def api_mobile_task_create():
    payload = request.get_json(force=True)
    task, token = create_mobile_task_from_payload(payload)
    return jsonify(mobile_task_response(task, token))


@app.route("/api/mobile/task/<task_id>/qrcode")
def api_mobile_task_qrcode(task_id):
    task = load_mobile_task(task_id)
    try:
        import qrcode
    except ImportError:
        raise UserFacingError("未安装 qrcode，请执行 pip install qrcode[pil]。", 500)

    image = qrcode.make(task.get("mobile_url") or "")
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return send_file(buffer, mimetype="image/png", download_name="%s.png" % safe_name(task_id, "mobile_qr"))


@app.route("/api/mobile/task")
def api_mobile_task():
    task_id = request.args.get("task_id") or ""
    token = request.args.get("token") or ""
    task = load_mobile_task(task_id)
    assert_mobile_task_token(task, token)
    task["last_seen_at"] = iso_now()
    save_mobile_task(task)
    return jsonify(
        {
            "ok": True,
            "task_id": task["task_id"],
            "session_id": task["session_id"],
            "base_info": task.get("base_info") or {},
            "items": read_excel_template(),
            "item_data": task.get("items") or [],
            "updated_seq": task.get("updated_seq") or 0,
            "expires_at": task.get("expires_at") or "",
        }
    )


@app.route("/api/mobile/item/save", methods=["POST"])
def api_mobile_item_save():
    payload = request.get_json(force=True)
    task = load_mobile_task(payload.get("task_id") or "")
    assert_mobile_task_token(task, payload.get("token") or "")
    item_id = normalize_text(payload.get("item_id"))
    if not item_id:
        raise UserFacingError("未指定排查项。", 400)

    item = find_payload_item(task, item_id)
    if "measured_value" in payload:
        item["measured_value"] = normalize_text(payload.get("measured_value"))
    if "conclusion" in payload:
        item["conclusion"] = normalize_conclusion(payload.get("conclusion"))
    item["mobile_updated_at"] = iso_now()
    task["last_seen_at"] = iso_now()
    task["updated_at"] = iso_now()
    task["updated_seq"] = int(task.get("updated_seq") or 0) + 1
    save_mobile_task(task)
    return jsonify({"ok": True, "item": item, "updated_seq": task["updated_seq"]})


@app.route("/api/mobile/upload", methods=["POST"])
def api_mobile_upload():
    task = load_mobile_task(request.form.get("task_id") or "")
    assert_mobile_task_token(task, request.form.get("token") or "")
    item_id = normalize_text(request.form.get("item_id"))
    field = normalize_text(request.form.get("field"))
    item = validate_mobile_upload_context(task, item_id, field)
    current_count = len(item.get(field) or [])
    incoming_count = len(request.files.getlist("files"))
    if current_count + incoming_count > MAX_IMAGES_PER_FIELD:
        raise UserFacingError("每个位置最多上传 %s 张照片，请删除多余图片后继续。" % MAX_IMAGES_PER_FIELD)

    session_id, uploaded = save_uploaded_files(task.get("session_id"), request.files.getlist("files"))
    task["session_id"] = session_id
    append_unique_images(item, field, uploaded)
    item["mobile_updated_at"] = iso_now()
    task["last_seen_at"] = iso_now()
    task["updated_at"] = iso_now()
    task["updated_seq"] = int(task.get("updated_seq") or 0) + 1
    save_mobile_task(task)
    return jsonify({"ok": True, "session_id": session_id, "files": uploaded, "item": item, "updated_seq": task["updated_seq"]})


@app.route("/api/mobile/upload/chunk/status", methods=["POST"])
def api_mobile_upload_chunk_status():
    payload = request.get_json(force=True)
    task = load_mobile_task(payload.get("task_id") or "")
    assert_mobile_task_token(task, payload.get("token") or "")
    upload_id = normalize_text(payload.get("upload_id"))
    if not upload_id:
        raise UserFacingError("上传任务无效。", 400)
    metadata = load_chunk_metadata(task.get("task_id"), upload_id)
    return jsonify(
        {
            "ok": True,
            "upload_id": upload_id,
            "uploaded_chunks": uploaded_chunk_indexes(task.get("task_id"), upload_id),
            "metadata": metadata,
            "chunk_size": MOBILE_CHUNK_SIZE,
        }
    )


@app.route("/api/mobile/upload/chunk", methods=["POST"])
def api_mobile_upload_chunk():
    task = load_mobile_task(request.form.get("task_id") or "")
    assert_mobile_task_token(task, request.form.get("token") or "")
    task_id = task.get("task_id")
    item_id = normalize_text(request.form.get("item_id"))
    field = normalize_text(request.form.get("field"))
    upload_id = normalize_text(request.form.get("upload_id"))
    validate_mobile_upload_context(task, item_id, field)
    if not upload_id:
        raise UserFacingError("上传任务无效。", 400)

    chunk_index = parse_positive_int(request.form.get("chunk_index"), "分片序号", minimum=0)
    chunk_count = parse_positive_int(request.form.get("chunk_count"), "分片总数", minimum=1, maximum=10000)
    original_size = parse_positive_int(request.form.get("original_size") or 1, "原始大小", minimum=1)
    processed_size = parse_positive_int(request.form.get("processed_size") or original_size, "处理后大小", minimum=1)
    if chunk_index >= chunk_count:
        raise UserFacingError("分片序号超出范围。", 400)

    chunk_file = request.files.get("chunk")
    if not chunk_file:
        raise UserFacingError("未收到图片分片。", 400)

    original_name = normalize_text(request.form.get("file_name")) or "mobile_photo.jpg"
    if not allowed_file(original_name):
        original_name = safe_name(Path(original_name).stem or "mobile_photo", "mobile_photo") + ".jpg"

    upload_dir = mobile_chunk_upload_dir(task_id, upload_id)
    chunks_dir = upload_dir / "chunks"
    chunks_dir.mkdir(parents=True, exist_ok=True)
    metadata = load_chunk_metadata(task_id, upload_id)
    if metadata:
        if metadata.get("item_id") != item_id or metadata.get("field") != field:
            raise UserFacingError("上传任务与排查项不匹配，请重新选择照片。", 400)
        if int(metadata.get("chunk_count") or chunk_count) != chunk_count:
            raise UserFacingError("上传分片信息不一致，请重新选择照片。", 400)
    else:
        metadata = {
            "task_id": task_id,
            "upload_id": upload_id,
            "item_id": item_id,
            "field": field,
            "file_name": original_name,
            "original_size": original_size,
            "processed_size": processed_size,
            "chunk_count": chunk_count,
            "created_at": iso_now(),
            "client_meta": normalize_text(request.form.get("client_meta")),
        }

    part_path = chunks_dir / ("%06d.part" % chunk_index)
    chunk_file.save(str(part_path))
    metadata["updated_at"] = iso_now()
    save_chunk_metadata(task_id, upload_id, metadata)

    uploaded = uploaded_chunk_indexes(task_id, upload_id)
    task["last_seen_at"] = iso_now()
    save_mobile_task(task)
    return jsonify(
        {
            "ok": True,
            "upload_id": upload_id,
            "uploaded_chunks": uploaded,
            "received": chunk_index,
            "complete": len(uploaded) >= chunk_count,
        }
    )


@app.route("/api/mobile/upload/complete", methods=["POST"])
def api_mobile_upload_complete():
    payload = request.get_json(force=True)
    task = load_mobile_task(payload.get("task_id") or "")
    assert_mobile_task_token(task, payload.get("token") or "")
    task_id = task.get("task_id")
    upload_id = normalize_text(payload.get("upload_id"))
    if not upload_id:
        raise UserFacingError("上传任务无效。", 400)

    metadata = load_chunk_metadata(task_id, upload_id)
    if not metadata:
        raise UserFacingError("未找到待合并照片，请重新上传。", 404)

    item_id = normalize_text(metadata.get("item_id"))
    field = normalize_text(metadata.get("field"))
    item = validate_mobile_upload_context(task, item_id, field)
    current_count = len(item.get(field) or [])
    if current_count >= MAX_IMAGES_PER_FIELD:
        raise UserFacingError("每个位置最多上传 %s 张照片，请删除多余图片后继续。" % MAX_IMAGES_PER_FIELD)

    chunk_count = parse_positive_int(metadata.get("chunk_count"), "分片总数", minimum=1, maximum=10000)
    uploaded = uploaded_chunk_indexes(task_id, upload_id)
    missing = [index for index in range(chunk_count) if index not in uploaded]
    if missing:
        return jsonify({"ok": False, "message": "照片分片未传完整。", "missing_chunks": missing[:50]}), 409

    upload_dir = mobile_chunk_upload_dir(task_id, upload_id)
    merged_name = safe_name(Path(metadata.get("file_name") or "mobile_photo.jpg").stem, "mobile_photo") + "." + (
        get_extension(metadata.get("file_name") or "") or "jpg"
    )
    merged_path = upload_dir / merged_name
    with merged_path.open("wb") as target:
        for index in range(chunk_count):
            part_path = upload_dir / "chunks" / ("%06d.part" % index)
            with part_path.open("rb") as source:
                shutil.copyfileobj(source, target)

    client_meta = {}
    if metadata.get("client_meta"):
        try:
            client_meta = json.loads(metadata.get("client_meta") or "{}")
        except json.JSONDecodeError:
            client_meta = {}

    session_id, image = register_uploaded_image(
        task.get("session_id"),
        merged_path,
        metadata.get("file_name") or "mobile_photo.jpg",
        original_size=metadata.get("original_size"),
        extra={
            "source": "mobile",
            "upload_id": upload_id,
            "watermarked": bool(client_meta.get("watermarked")),
            "client_compressed": bool(client_meta.get("compressed")),
            "client_original_size": client_meta.get("original_size") or metadata.get("original_size"),
            "client_processed_size": client_meta.get("processed_size") or metadata.get("processed_size"),
            "watermark_text": client_meta.get("watermark_text") or "",
        },
    )
    task["session_id"] = session_id
    append_unique_images(item, field, [image])
    item["mobile_updated_at"] = iso_now()
    task["last_seen_at"] = iso_now()
    task["updated_at"] = iso_now()
    task["updated_seq"] = int(task.get("updated_seq") or 0) + 1
    save_mobile_task(task)
    shutil.rmtree(str(upload_dir), ignore_errors=True)
    return jsonify({"ok": True, "session_id": session_id, "file": image, "item": item, "updated_seq": task["updated_seq"]})


@app.route("/api/mobile/task/status")
def api_mobile_task_status():
    task = load_mobile_task(request.args.get("task_id") or "")
    assert_mobile_task_token(task, request.args.get("token") or "")
    return jsonify(
        {
            "ok": True,
            "task_id": task.get("task_id"),
            "session_id": task.get("session_id"),
            "status": task.get("status"),
            "connected": bool(task.get("last_seen_at")),
            "last_seen_at": task.get("last_seen_at") or "",
            "updated_at": task.get("updated_at") or "",
            "updated_seq": task.get("updated_seq") or 0,
            "base_info": task.get("base_info") or {},
            "items": task.get("items") or [],
        }
    )


# -----------------------------
# RTS 审核返回报告模块 V1.2
# -----------------------------

RTS_REVIEW_FIELDS = {
    "review_conclusion": "审核结论",
    "initial_judgement": "初步判断",
    "supplement_required": "是否需要补充资料",
    "supplement_items": "需补充资料",
    "suggestions": "下一步处理建议",
    "upgrade_required": "是否需要升级",
    "upgrade_target": "升级方向",
    "reviewer": "审核人",
    "review_date": "审核时间",
    "review_notes": "审核说明",
}


def load_report_data_from_zip(zip_path, session_id):
    """Extract a front-line report ZIP safely and return report_data.json data plus session root."""
    extract_root = RTS_UPLOADS_DIR / session_id / "extracted"
    extract_root.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(str(zip_path), "r") as archive:
        json_member = None
        for member in archive.infolist():
            member_name = member.filename.replace("\\", "/")
            if member_name.endswith("/"):
                continue
            if ".." in Path(member_name).parts:
                raise UserFacingError("ZIP 文件包含异常路径，请检查后重新上传。")
            if Path(member_name).name == "report_data.json":
                json_member = member_name
            target = (extract_root / member_name).resolve()
            if extract_root.resolve() != target and extract_root.resolve() not in target.parents:
                raise UserFacingError("ZIP 文件包含异常路径，请检查后重新上传。")
            target.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(member) as src, open(target, "wb") as dst:
                shutil.copyfileobj(src, dst)

    if not json_member:
        raise UserFacingError("ZIP 报告包中未找到 report_data.json，请确认上传的是一线报告 ZIP。")

    data_path = extract_root / json_member
    data = json.loads(data_path.read_text(encoding="utf-8"))
    return data, extract_root


def load_report_data_from_json(file_storage, session_id):
    upload_root = RTS_UPLOADS_DIR / session_id
    upload_root.mkdir(parents=True, exist_ok=True)
    json_path = upload_root / "report_data.json"
    file_storage.save(str(json_path))
    data = json.loads(json_path.read_text(encoding="utf-8"))
    return data, upload_root


def load_rts_review_data_from_zip(zip_path, session_id):
    """Extract an RTS return ZIP safely and return rts_review_data.json data."""
    extract_root = RTS_UPLOADS_DIR / session_id / "rts_extracted"
    extract_root.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(str(zip_path), "r") as archive:
        json_member = None
        for member in archive.infolist():
            member_name = member.filename.replace("\\", "/")
            if member_name.endswith("/"):
                continue
            if ".." in Path(member_name).parts:
                raise UserFacingError("ZIP 文件包含异常路径，请检查后重新上传。")
            if Path(member_name).name == "rts_review_data.json":
                json_member = member_name
            target = (extract_root / member_name).resolve()
            if extract_root.resolve() != target and extract_root.resolve() not in target.parents:
                raise UserFacingError("ZIP 文件包含异常路径，请检查后重新上传。")
            target.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(member) as src, open(target, "wb") as dst:
                shutil.copyfileobj(src, dst)

    if not json_member:
        raise UserFacingError("RTS 返回 ZIP 中未找到 rts_review_data.json，请确认上传的是 RTS 审核返回 ZIP。")

    data_path = extract_root / json_member
    data = json.loads(data_path.read_text(encoding="utf-8"))
    return data, extract_root


def load_rts_review_data_from_json(file_storage, session_id):
    upload_root = RTS_UPLOADS_DIR / session_id
    upload_root.mkdir(parents=True, exist_ok=True)
    json_path = upload_root / "rts_review_data.json"
    file_storage.save(str(json_path))
    data = json.loads(json_path.read_text(encoding="utf-8"))
    return data, upload_root


def restore_editable_report_from_source(source_data, extract_root):
    if not source_data.get("groups"):
        raise UserFacingError("报告数据不完整，未找到排查项。")

    session_id = "restore_" + uuid.uuid4().hex
    warnings = []
    restored_items = []
    extract_root = Path(extract_root).resolve()

    for item in iter_report_items(source_data):
        restored = {
            "id": item.get("id"),
            "measured_value": normalize_text(item.get("measured_value")),
            "conclusion": normalize_conclusion(item.get("conclusion")),
            "before_images": [],
            "after_images": [],
        }
        if not restored["id"]:
            continue

        for field in ["before_images", "after_images"]:
            for image in item.get(field) or []:
                rel_path = (image.get("path") or "").replace("\\", "/")
                if not rel_path:
                    continue
                source = (extract_root / rel_path).resolve()
                if extract_root != source and extract_root not in source.parents:
                    warnings.append("%s 的图片路径异常，已跳过。" % (item.get("display_step") or item.get("step") or restored["id"]))
                    continue
                if not source.exists() or not source.is_file():
                    warnings.append("%s 的图片未在 ZIP 中找到：%s" % ((item.get("display_step") or item.get("step") or restored["id"]), rel_path))
                    continue
                try:
                    restored[field].append(cache_existing_report_image(session_id, source, image))
                except UserFacingError as exc:
                    warnings.append("%s 的图片恢复失败：%s" % ((item.get("display_step") or item.get("step") or restored["id"]), exc.message))

        restored_items.append(restored)

    if not restored_items:
        raise UserFacingError("报告数据中未找到可恢复的排查项。")

    base_info = dict(source_data.get("base_info") or {})
    return {
        "session_id": session_id,
        "base_info": base_info,
        "items": restored_items,
        "issue_no": source_data.get("issue_no") or base_info.get("issue_no") or "",
        "stats": source_data.get("stats") or {},
        "warnings": warnings,
    }


def iter_report_items(source_data):
    for group in source_data.get("groups") or []:
        for item in group.get("items") or []:
            yield item


def source_item_lookup(source_data):
    lookup = {}
    for item in iter_report_items(source_data or {}):
        keys = [item.get("id"), item.get("display_step"), item.get("step")]
        display_step = normalize_text(item.get("display_step") or item.get("step"))
        child_match = re.search(r"第\s*(\d+)\s*步-(\d+)", display_step)
        if child_match:
            step_no = int(child_match.group(1))
            child_no = int(child_match.group(2))
            keys.append("item_%03d_%s" % (step_no, child_no))
        for key in keys:
            text = normalize_text(key)
            if text:
                lookup[text] = item
    return lookup


def truthy(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return normalize_text(value).lower() in {"1", "true", "yes", "y", "是", "需要", "必补", "补充"}


def normalize_supplement_requests(requests, source_data=None):
    lookup = source_item_lookup(source_data or {})
    normalized = []
    seen = set()

    for request_item in requests or []:
        if not isinstance(request_item, dict):
            continue

        item_id = normalize_text(request_item.get("item_id") or request_item.get("id") or request_item.get("item_key"))
        item = lookup.get(item_id)
        if not item:
            display_key = normalize_text(request_item.get("display_step") or request_item.get("step"))
            item = lookup.get(display_key)

        if item:
            item_id = normalize_text(item.get("id")) or item_id
        if not item_id:
            continue

        fields = request_item.get("fields") or []
        if isinstance(fields, str):
            fields = [fields]
        field_set = {normalize_text(field) for field in fields}
        need_record = truthy(request_item.get("need_record")) or bool({"record", "measured_value", "实测情况记录"} & field_set)
        need_before = truthy(request_item.get("need_before")) or bool({"before", "before_images", "原始状态照片"} & field_set)
        need_after = truthy(request_item.get("need_after")) or bool({"after", "after_images", "调试或维护后照片"} & field_set)
        requirement = normalize_text(
            request_item.get("requirement")
            or request_item.get("suggestion")
            or request_item.get("notes")
            or request_item.get("text")
        )
        if not requirement:
            requirement = "请按 RTS 意见补充该项资料。"

        key = (item_id, need_record, need_before, need_after, requirement)
        if key in seen:
            continue
        seen.add(key)

        normalized.append(
            {
                "item_id": item_id,
                "display_step": normalize_text((item or {}).get("display_step") or (item or {}).get("step") or request_item.get("display_step")),
                "action": normalize_text((item or {}).get("action") or request_item.get("action")),
                "category": normalize_text((item or {}).get("category") or request_item.get("category")),
                "need_record": need_record,
                "need_before": need_before,
                "need_after": need_after,
                "requirement": requirement,
            }
        )

    return normalized


def build_attention_from_source(source_data):
    attention = source_data.get("attention_items") or []
    if attention:
        return attention
    return [item for item in iter_report_items(source_data) if normalize_conclusion(item.get("conclusion")) in {"异常", "已处理", "待确认"}]


def enrich_source_image_urls(source_data, session_id):
    """Add preview URLs for images extracted from ZIP. JSON-only uploads simply have no preview URL."""
    for item in iter_report_items(source_data):
        for field in ["before_images", "after_images"]:
            for image in item.get(field) or []:
                rel_path = (image.get("path") or "").replace("\\", "/")
                if rel_path:
                    image["preview_url"] = "/rts-source/%s/%s" % (session_id, rel_path)
    for item in source_data.get("attention_items") or []:
        for field in ["before_images", "after_images"]:
            for image in item.get(field) or []:
                rel_path = (image.get("path") or "").replace("\\", "/")
                if rel_path:
                    image["preview_url"] = "/rts-source/%s/%s" % (session_id, rel_path)
    return source_data


def copy_rts_source_images(source_data, source_session_id, output_dir):
    """Copy images from extracted front-line ZIP into the RTS report folder and rewrite paths."""
    if not source_session_id:
        return source_data
    extract_root = RTS_UPLOADS_DIR / source_session_id / "extracted"
    if not extract_root.exists():
        return source_data
    image_output = output_dir / "source_images"
    copied = {}

    def rewrite_image(image):
        rel = (image.get("path") or "").replace("\\", "/")
        if not rel:
            return image
        source = (extract_root / rel).resolve()
        if extract_root.resolve() != source and extract_root.resolve() not in source.parents:
            return image
        if not source.exists() or not source.is_file():
            return image
        if rel not in copied:
            image_output.mkdir(parents=True, exist_ok=True)
            target_name = safe_name(Path(rel).name, "image")
            target = image_output / target_name
            if target.exists():
                target = image_output / (Path(target_name).stem + "_" + uuid.uuid4().hex[:6] + Path(target_name).suffix)
            shutil.copy2(str(source), str(target))
            copied[rel] = "source_images/" + target.name
        image["path"] = copied[rel]
        return image

    for item in iter_report_items(source_data):
        for field in ["before_images", "after_images"]:
            item[field] = [rewrite_image(dict(img)) for img in (item.get(field) or [])]

    # report_data.json 中通常已经带有 attention_items；这里也要同步重写图片路径，
    # 否则 RTS 返回报告里的重点项仍可能引用原一线报告的 images/ 相对路径。
    rewritten_attention = []
    for item in source_data.get("attention_items") or []:
        new_item = dict(item)
        for field in ["before_images", "after_images"]:
            new_item[field] = [rewrite_image(dict(img)) for img in (new_item.get(field) or [])]
        rewritten_attention.append(new_item)
    if rewritten_attention:
        source_data["attention_items"] = rewritten_attention
    else:
        source_data["attention_items"] = [
            item for item in iter_report_items(source_data)
            if normalize_conclusion(item.get("conclusion")) in {"异常", "已处理", "待确认"}
        ]
    return source_data


def make_rts_review_no(source_issue_no):
    base = safe_name(source_issue_no, datetime.now().strftime("%Y%m%d"))
    return "RTS-%s-%s" % (base, datetime.now().strftime("%H%M%S"))


def make_rts_slug(source_data, review_no):
    base_info = source_data.get("base_info") or {}
    hospital = safe_name(base_info.get("hospital"), "医院")
    serial = safe_name(base_info.get("serial"), "设备序列号")
    return "RTS审核返回报告_%s_%s_%s" % (hospital, serial, safe_name(review_no, "RTS"))


def validate_rts_review_payload(payload):
    errors = []
    review = payload.get("review") or {}
    source_data = payload.get("source_data") or {}
    supplement_requests = normalize_supplement_requests(payload.get("supplement_requests") or [], source_data)
    if not source_data.get("base_info"):
        errors.append("未读取到一线报告基础信息，请先上传一线报告包或 report_data.json。")
    for key in ["review_conclusion", "review_notes", "review_date"]:
        if not normalize_text(review.get(key)):
            errors.append("%s为必填。" % RTS_REVIEW_FIELDS[key])
    if normalize_text(review.get("supplement_required")) == "是" and not normalize_text(review.get("supplement_items")) and not supplement_requests:
        errors.append("已选择需要补充资料，请填写文字说明或勾选结构化补充清单。")
    return errors


def summarize_review_tags(review):
    judgement = review.get("initial_judgement") or []
    if isinstance(judgement, str):
        judgement = [judgement] if judgement else []
    return {
        "initial_judgement_text": "、".join(judgement) if judgement else "待确认",
        "upgrade_text": normalize_text(review.get("upgrade_target")) if normalize_text(review.get("upgrade_required")) == "是" else "暂不升级",
    }


@app.route("/rts")
def rts_review_page():
    return render_template("rts_review.html", app_name=APP_NAME, version=VERSION)


@app.route("/api/rts/import", methods=["POST"])
def api_rts_import():
    file_storage = request.files.get("report_file")
    if not file_storage or not file_storage.filename:
        raise UserFacingError("请选择一线报告 ZIP 或 report_data.json。")
    original_name = file_storage.filename
    session_id = "rts_" + uuid.uuid4().hex
    session_root = RTS_UPLOADS_DIR / session_id
    session_root.mkdir(parents=True, exist_ok=True)

    lower = original_name.lower()
    if lower.endswith(".zip"):
        zip_path = session_root / "source_report.zip"
        file_storage.save(str(zip_path))
        source_data, _ = load_report_data_from_zip(zip_path, session_id)
    elif lower.endswith(".json"):
        source_data, _ = load_report_data_from_json(file_storage, session_id)
    else:
        raise UserFacingError("文件格式不支持，请上传一线报告 ZIP 或 report_data.json。")

    source_data = enrich_source_image_urls(source_data, session_id)
    source_data["attention_items"] = build_attention_from_source(source_data)
    logging.info("RTS source imported: session=%s issue=%s", session_id, source_data.get("issue_no"))
    return jsonify({"ok": True, "session_id": session_id, "source_data": source_data})


@app.route("/api/rts/report", methods=["POST"])
def api_rts_report():
    payload = request.get_json(force=True)
    errors = validate_rts_review_payload(payload)
    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    source_data = payload.get("source_data") or {}
    review = dict(payload.get("review") or {})
    item_reviews = payload.get("item_reviews") or {}
    supplement_requests = normalize_supplement_requests(payload.get("supplement_requests") or [], source_data)
    source_session_id = payload.get("source_session_id") or ""

    review["supplement_required"] = "是" if supplement_requests else "否"
    if supplement_requests and not normalize_text(review.get("supplement_items")):
        review["supplement_items"] = "请按结构化补充清单返回补充资料。"
    review["suggestions"] = normalize_text(review.get("suggestions")) or normalize_text(review.get("review_notes")) or "请按 RTS 审核说明执行。"
    review["upgrade_required"] = normalize_text(review.get("upgrade_required")) or "否"
    review["upgrade_target"] = normalize_text(review.get("upgrade_target"))
    review["reviewer"] = normalize_text(review.get("reviewer")) or "RTS"
    review["initial_judgement"] = review.get("initial_judgement") or []

    source_issue_no = source_data.get("issue_no") or (source_data.get("base_info") or {}).get("issue_no") or "未编号"
    review_no = make_rts_review_no(source_issue_no)
    review_tags = summarize_review_tags(review)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    slug = make_rts_slug(source_data, review_no)
    output_dir = unique_output_dir(slug)
    output_dir.mkdir(parents=True, exist_ok=True)

    source_data_for_report = json.loads(json.dumps(source_data, ensure_ascii=False))
    source_data_for_report = copy_rts_source_images(source_data_for_report, source_session_id, output_dir)
    attention_items = build_attention_from_source(source_data_for_report)

    html = render_template(
        "rts_report.html",
        app_name=APP_NAME,
        version=VERSION,
        source_data=source_data_for_report,
        source_issue_no=source_issue_no,
        review_no=review_no,
        review=review,
        review_tags=review_tags,
        item_reviews=item_reviews,
        attention_items=attention_items,
        supplement_requests=supplement_requests,
        generated_at=generated_at,
    )

    report_path = output_dir / (slug + ".html")
    data_path = output_dir / "rts_review_data.json"
    zip_path = output_dir / (slug + ".zip")
    report_path.write_text(html, encoding="utf-8")
    data_path.write_text(json.dumps({
        "review_no": review_no,
        "source_issue_no": source_issue_no,
        "review": review,
        "review_tags": review_tags,
        "item_reviews": item_reviews,
        "supplement_requests": supplement_requests,
        "source_data": source_data_for_report,
        "attention_items": attention_items,
        "generated_at": generated_at,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    build_zip(output_dir, zip_path)
    logging.info("RTS review report generated: %s", report_path)
    return jsonify({
        "ok": True,
        "review_no": review_no,
        "output_dir": str(output_dir),
        "report_file": str(report_path),
        "zip_file": str(zip_path),
        "report_url": make_public_output_path(report_path),
        "zip_url": make_public_output_path(zip_path),
    })


@app.route("/rts-source/<path:session>/<path:filename>")
def rts_source_file(session, filename):
    root = (RTS_UPLOADS_DIR / session / "extracted").resolve()
    if not root.exists():
        # JSON-only upload: no extracted image root.
        root = (RTS_UPLOADS_DIR / session).resolve()
    return send_from_directory(str(root), filename)


@app.route("/api/drafts", methods=["GET"])
def api_drafts():
    drafts = []
    for path in sorted(DRAFTS_DIR.glob("*.json"), key=lambda item: item.stat().st_mtime, reverse=True):
        drafts.append(
            {
                "name": path.name,
                "updated_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "size": path.stat().st_size,
            }
        )
    return jsonify({"ok": True, "drafts": drafts})


@app.route("/api/draft/save", methods=["POST"])
def api_draft_save():
    payload = request.get_json(force=True)
    base_info = payload.get("base_info") or {}
    filename = "草稿_%s_%s_%s.json" % (
        safe_name(base_info.get("hospital"), "医院"),
        safe_name(base_info.get("serial"), "设备序列号"),
        datetime.now().strftime("%Y%m%d_%H%M%S"),
    )
    path = DRAFTS_DIR / filename
    payload["saved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    logging.info("draft saved: %s", path)
    return jsonify({"ok": True, "name": filename, "path": str(path)})


@app.route("/api/draft/load/<path:name>")
def api_draft_load(name):
    path = (DRAFTS_DIR / name).resolve()
    root = DRAFTS_DIR.resolve()
    if root != path and root not in path.parents:
        raise UserFacingError("草稿路径异常。")
    if not path.exists():
        raise UserFacingError("草稿文件不存在。", 404)
    return jsonify({"ok": True, "draft": json.loads(path.read_text(encoding="utf-8"))})


@app.route("/api/open-output", methods=["POST"])
def api_open_output():
    payload = request.get_json(force=True)
    path = Path(payload.get("path") or OUTPUT_DIR).resolve()
    root = OUTPUT_DIR.resolve()
    if root != path and root not in path.parents:
        raise UserFacingError("输出路径异常。")
    if not path.exists():
        raise UserFacingError("输出目录不存在。", 404)
    if os.name == "nt":
        os.startfile(str(path))
    else:
        webbrowser.open(str(path))
    return jsonify({"ok": True})


@app.route("/uploads/<path:session>/<path:filename>")
def uploaded_file(session, filename):
    return send_from_directory(str(UPLOADS_DIR / session), filename)


@app.route("/output-files/<path:filename>")
def output_file(filename):
    return send_from_directory(str(OUTPUT_DIR), filename)


def find_free_port(start=5000, end=5100):
    for port in range(start, end + 1):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            try:
                sock.bind(("127.0.0.1", port))
                return port
            except OSError:
                continue
    raise RuntimeError("No free local port found.")


def open_browser_later(port):
    def open_page():
        webbrowser.open("http://127.0.0.1:%s" % port)

    threading.Timer(1.0, open_page).start()


def main():
    setup_runtime()
    port = int(os.environ.get("JUMP_CHECK_PORT") or find_free_port())
    open_browser_later(port)
    app.run(host="0.0.0.0", port=port, debug=False)


if __name__ == "__main__":
    main()
