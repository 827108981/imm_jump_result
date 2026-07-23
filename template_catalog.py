"""Latest frontline template catalogue and reference-image metadata.

The workbook is the single source of truth for the step wording, record examples,
and image anchors.  The UI decides visibility from the explicit condition fields
returned with every item; the report API applies the same rules server-side.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_NAME = "M1武汉市第八医院-双大客户仪器故障反馈和排查表-20250412.xlsx"
SHEET_NAME = "附件1-跳高值排查指南"
REFERENCE_DIR_NAME = "reference"

MODELS = ["CL-6000i", "CL-8000i"]
ULTRASOUND_PROJECTS = [
    "PROG", "CEA", "TSH", "HBsAg", "HBsAg-I", "HBsAg II", "HBeAg", "HIV", "TnI",
    "CYFRA 21-1", "Anti-HCV", "Anti-TP", "Anti-TP II", "HBeAg Q", "PCT", "Tβ HCG-I",
    "D TβHCG-II", "TNI(V1.02.01)", "E2(V1.03.01)", "PTH(V1.01.T)", "CK-MB(V1.01T)",
    "MYO (V1.01T)",
]


def normalize_text(value):
    return "" if value is None else str(value).replace("\u3000", " ").strip()


def is_na(value):
    return normalize_text(value).upper().replace(" ", "") in {"N/A", "NA", "N／A"}


def normalize_project(value):
    text = normalize_text(value).upper()
    text = text.replace("Ｉ", "I").replace("Ⅱ", "II").replace("Β", "B")
    return re.sub(r"[\s\-_（）()]", "", text)


ULTRASOUND_PROJECT_KEYS = {normalize_project(project) for project in ULTRASOUND_PROJECTS}

REFERENCE_IMAGE_TARGET_ROWS = {
    "image92.jpg": 17,
    "image93.jpg": 17,
    "image94.jpg": 17,
    "image43.jpg": 31,
    "image44.jpg": 31,
    "image25.jpg": 38,
}


def normalize_record_example(value):
    """Apply approved corrections before a workbook example reaches either UI."""
    text = normalize_text(value)
    return re.sub(
        r"(\u4eea\u5668\u63a5\u5934\u7aef\u6d4b\u96f6\u5730\u7535\u538b\s*[\uff1a:]\s*)219\s*V",
        r"\g<1>0.8 V",
        text,
        flags=re.IGNORECASE,
    )


def _example_field(label, value):
    return {"label": normalize_text(label), "value": normalize_text(value)}


def _append_example_fields(blocks, fields, title=""):
    clean_fields = [field for field in fields if field["label"] and field["value"]]
    if clean_fields:
        blocks.append({"type": "fields", "title": normalize_text(title), "fields": clean_fields})


def format_record_example(value):
    """Convert spacing-dependent Excel examples into portable field blocks.

    The Excel template contains visual spacing meant for wide cells.  Returning a
    compact structure lets desktop and phone views align the same information
    without relying on those spaces.
    """
    text = normalize_record_example(value)
    if not text:
        return []

    if "返回性能测试数据文件" in text and "CL6000i目录" in text and "CL2000i" in text:
        fields = []
        notes = []
        for raw_line in text.replace("\r", "").split("\n"):
            line = normalize_text(raw_line)
            if not line:
                continue
            if line.startswith("CL6000i目录"):
                fields.append(_example_field("CL-6000i目录", re.sub(r"^CL6000i目录\s*[：:]\s*", "", line)))
            elif line.startswith("CL2000i"):
                fields.append(_example_field("CL-2000i（1000i）目录", re.sub(r"^CL2000i.*?目录\s*[：:]\s*", "", line)))
            elif "返回性能测试数据文件" not in line:
                notes.append({"type": "note", "text": line})
        if fields:
            return notes + [{"type": "fields", "title": "返回性能测试数据文件", "layout": "stacked", "fields": fields}]

    compact = re.sub(r"[ \t]+", " ", text.replace("\r", ""))
    if "\u78c1\u5206\u79bb\u76d8" in compact and "\u6db2\u6d41\u91cf" in compact:
        blocks = []
        marker_pattern = re.compile(r"\u78c1\u5206\u79bb\u76d8\s*(\d+)\s*[\uff1a:]")
        markers = list(marker_pattern.finditer(compact))
        for index, marker in enumerate(markers):
            section = compact[marker.end(): markers[index + 1].start() if index + 1 < len(markers) else len(compact)]
            fields = [
                _example_field(label, reading)
                for label, reading in re.findall(r"(\u7b2c[\u4e00\u4e8c\u4e09\u56db]\u9636\u6db2\u6d41\u91cf)\s*[\uff1a:]\s*([0-9.]+)", section)
            ]
            _append_example_fields(blocks, fields, "\u78c1\u5206\u79bb\u76d8%s" % marker.group(1))
        if blocks:
            return blocks

    flow_match = re.search(
        r"(\u6837\u672c\u76ee\u6807\u6d41\u91cf)\s*[\uff1a:]\s*([^\s;\uff1b]+).*?"
        r"(\u5b9e\u9645\u6d41\u91cf)\s*[\uff1a:]\s*([^\s;\uff1b]+).*?"
        r"(\u8bd5\u5242\u76ee\u6807\u6d41\u91cf)\s*[\uff1a:]\s*([^\s;\uff1b]+).*?"
        r"(\u5b9e\u9645\u6d41\u91cf)\s*[\uff1a:]\s*([^\s;\uff1b]+)",
        compact,
        flags=re.DOTALL,
    )
    if flow_match:
        return [
            {"type": "fields", "title": "\u6837\u672c", "fields": [_example_field(flow_match.group(1), flow_match.group(2)), _example_field(flow_match.group(3), flow_match.group(4))]},
            {"type": "fields", "title": "\u8bd5\u5242", "fields": [_example_field(flow_match.group(5), flow_match.group(6)), _example_field(flow_match.group(7), flow_match.group(8))]},
        ]

    blocks = []
    fields = []
    section_title = ""

    def flush_fields():
        nonlocal fields, section_title
        _append_example_fields(blocks, fields, section_title)
        fields = []
        section_title = ""

    for raw_line in compact.split("\n"):
        line = normalize_text(raw_line)
        if not line:
            continue
        if re.fullmatch(r"[^\uff1a:]+[\uff1a:]", line):
            flush_fields()
            section_title = line[:-1]
            continue

        mechanical = re.match(r"((?:.*?)(?:\u6c34\u5e73|\u5782\u76f4|\u5b9a\u4f4d))\s+(.+)$", line)
        if not mechanical:
            mechanical = re.match(r"(\u5438\u6837\u5b9a\u4f4d)(.+)$", line)
        if mechanical:
            fields.append(_example_field(mechanical.group(1), mechanical.group(2)))
            continue

        pairs = re.findall(r"([^\uff1a:;\uff1b\n]+?)\s*[\uff1a:]\s*([^;\uff1b\n]+)", line)
        if pairs:
            for label, reading in pairs:
                fields.append(_example_field(label, reading))
            continue

        flush_fields()
        blocks.append({"type": "note", "text": line})

    flush_fields()
    return blocks


def normalize_model(value):
    text = normalize_text(value).upper().replace(" ", "")
    if "8000" in text:
        return "CL-8000i"
    if "6000" in text:
        return "CL-6000i"
    return normalize_text(value)


def is_ultrasound_context(base_info):
    base_info = base_info or {}
    return (
        normalize_model(base_info.get("model")) == "CL-8000i"
        and normalize_project(base_info.get("jump_project")) in ULTRASOUND_PROJECT_KEYS
    )


def item_is_applicable(item, base_info):
    condition = item.get("condition") or {}
    model = normalize_model((base_info or {}).get("model"))
    if condition.get("models") and model not in condition["models"]:
        return False
    if condition.get("ultrasound_only") and not is_ultrasound_context(base_info):
        return False
    return True


def display_step_for_context(item, base_info):
    if is_ultrasound_context(base_info):
        return item.get("display_step_special") or item.get("display_step") or ""
    return item.get("display_step_default") or item.get("display_step") or ""


def prepare_items_for_context(items, base_info):
    prepared = []
    for item in items:
        if not item_is_applicable(item, base_info):
            continue
        copy = dict(item)
        copy["display_step"] = display_step_for_context(copy, base_info)
        prepared.append(copy)
    return sorted(prepared, key=lambda item: item.get("sort_order", 0))


def get_project_options(workbook_path):
    workbook = load_workbook(str(workbook_path), data_only=True, read_only=True)
    values = []
    if len(workbook.worksheets) > 1:
        for row in workbook.worksheets[1].iter_rows(values_only=True):
            value = normalize_text(row[0] if row else "")
            if value and value not in values and value not in {"测试项目", "跳值项目", "项目名称"}:
                values.append(value)
    for project in ULTRASOUND_PROJECTS:
        if project not in values:
            values.append(project)
    return values


def _cell_text(sheet, row, column):
    return normalize_text(sheet.cell(row, column).value)


def _base_item(sheet, row, item_id, sort_order, display_default, display_special=None):
    record_example = normalize_record_example(_cell_text(sheet, row, 7))
    before_template = _cell_text(sheet, row, 8)
    after_template = _cell_text(sheet, row, 9)
    item = {
        "id": item_id,
        "index": sort_order // 100,
        "sort_order": sort_order,
        "step": display_default,
        "display_step": display_default,
        "display_step_default": display_default,
        "display_step_special": display_special or display_default,
        "priority": _cell_text(sheet, row, 2) or "未分类",
        "category": _cell_text(sheet, row, 3) or "未分类",
        "action": _cell_text(sheet, row, 4),
        "standard": _cell_text(sheet, row, 5),
        "execution": _cell_text(sheet, row, 6) or "是",
        "record_template": "",
        "record_example": "" if is_na(record_example) else record_example,
        "record_example_blocks": [] if is_na(record_example) else format_record_example(record_example),
        "before_template": before_template,
        "after_template": after_template,
        "record_required": not is_na(record_example),
        "before_required": not is_na(before_template),
        "after_required": not is_na(after_template),
        "reference_images": [],
    }
    if normalize_text(item["action"]) == "低值质控重复性测试":
        item["record_example"] = "测试200次，跳值率0.5%"
        item["record_example_blocks"] = [{"type": "note", "text": item["record_example"]}]
    return item


def apply_reference_image_overrides(references):
    """Move Excel figures whose anchors do not match their intended check step."""
    adjusted = {row: list(images) for row, images in references.items()}
    moved = []
    for row, images in adjusted.items():
        retained = []
        for image in images:
            target_row = REFERENCE_IMAGE_TARGET_ROWS.get(image.get("name"))
            if target_row:
                moved.append((target_row, image))
            else:
                retained.append(image)
        adjusted[row] = retained
    for target_row, image in moved:
        adjusted.setdefault(target_row, []).append(image)
    return adjusted


def build_template_items(workbook_path, reference_images=None):
    workbook = load_workbook(str(workbook_path), data_only=True)
    sheet = workbook[SHEET_NAME]
    references = apply_reference_image_overrides(reference_images or {})
    items = []

    for row in range(4, 38):
        step_no = row - 3
        item = _base_item(sheet, row, "item_%03d" % step_no, step_no * 100, "第%d步" % step_no)
        if step_no == 31:
            item["condition"] = {"models": ["CL-6000i"]}
        item["reference_images"] = list(references.get(row, []))
        items.append(item)

    for offset, row in enumerate(range(38, 45), 1):
        item = _base_item(
            sheet,
            row,
            "item_035_ultrasound_%d" % offset,
            3500 + offset,
            "第35步-%d" % offset,
        )
        item["index"] = 35
        item["condition"] = {"models": ["CL-8000i"], "ultrasound_only": True}
        item["record_required"] = True
        item["before_required"] = False
        item["after_required"] = False
        item["reference_images"] = list(references.get(row, []))
        items.append(item)

    for offset, row in enumerate(range(45, 49), 1):
        item_id = "item_035" if offset == 1 else "item_035_%d" % offset
        item = _base_item(
            sheet,
            row,
            item_id,
            3600 + offset,
            "第35步-%d" % offset,
            "第36步-%d" % offset,
        )
        item["index"] = 35
        if offset <= 3:
            item["record_required"] = False
            item["before_required"] = False
            item["after_required"] = True
        else:
            item["record_required"] = True
            item["before_required"] = False
            item["after_required"] = False
        item["reference_images"] = list(references.get(row, []))
        items.append(item)

    log_item = _base_item(sheet, 49, "item_036", 3700, "第36步", "第37步")
    log_item["index"] = 36
    log_item["reference_images"] = list(references.get(49, []))
    items.append(log_item)
    return items
